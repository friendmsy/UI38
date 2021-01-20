#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020/5/6 10:40
# @Author : msy
import openpyxl
import xlrd
import xlsxwriter as xlsxwriter
import xlwt
from xlutils.copy import copy
from common.log import *
# import yaml
from common.commethod import *
from xlwt import XFStyle
import threading
from openpyxl import load_workbook,Workbook
import shutil
import zipfile


class DataOperate():
    lock = threading.Lock()

    @staticmethod
    # file:用例具体路径，rerun：是根据是否重跑，要获取用例，默认不是重跑用例
    def get_case(file,rerun=None):
        book = xlrd.open_workbook(file)  # 创建一个excel操作对象
        # print(os.path.join(CASEPATH, file))
        sheet = book.sheet_by_index(0)  # 创建一个sheet操作实例,读取的是第一个excel，如果需要也可以参数化
        nnrow = sheet.nrows  # 获取行数
        # 创建一个空列表，用于存放用例
        all_case=[]
        # print('空列表长度',len(all_case))

        for j in range(2, nnrow):   #不是从0开始读取，从第三行开始读取
            if  rerun is None and sheet.row_values(j)[2]!='#':  #  '#'表示此行跳过，不获取，第三列valid
                list1=list(sheet.row_values(j)[3:15])
                list1.append(j)
                # 把行数也加入到列表，目的是写入时方便定位到对应的行数;指定行指定列数据，从valid开始到row
                all_case.append(list1)    #append里面不能再调用append，会导致返回的是None
            elif rerun is not None and sheet.row_values(j)[2]!='#' and sheet.row_values(j)[12]=='0':
                print('进入重跑获取用例中.......')
                # print(sheet.row_values(j)[12])
                list1 = list(sheet.row_values(j)[3:15])
                list1.append(j)
                # 把行数也加入到列表，目的是写入时方便定位到对应的行数;指定行指定列数据，从valid开始到row
                all_case.append(list1)  # append里面不能再调用append，会导致返回的是None
            else:
                # print('没有可获取的用例！')
                # 空列表，用于以下判断是否有获取到用例
                pass

        if len(all_case)==0:
            # 返回1，目的是标记成没有获取到用例，直接返回结果1，不执行用例，直接通过
            all_cases=['1']
        else:
            all_cases=all_case
        # Mylog().my_log().info('i遍历获取用例')

        return all_cases

    '''
    def copy_excel(self,dirpath,filepath,filepathnext='备份'):
        # 在当前目录下创建一个新的目录用于存入复制得到的EXCEL
        newfilepath=os.path.join(dirpath,filepathnext)
        # 创建目录
        if not os.path.exists(newfilepath):
            os.mkdir(newfilepath)
            print('输出创建的目录', newfilepath)
        else:
            print('不需要创建访目录')
        # 从原文件中读excel内容
        book=xlrd.open_workbook(filepath)
        sheet=book.sheet_by_index(0)
        nrows=sheet.nrows
        cols=sheet.ncols
        content=[]
        for i in range(nrows):
            # list1=list(sheet.row_values(i)[0:20])
            content.append(list(sheet.row_values(i)))
        print(content)
        # 把原文件名取出来，以用于创建副件名
        copyexcel=os.path.join(newfilepath,os.path.basename(filepath).split('.')[0]+'_副本'+'.xlsx')
        # 创建空excel,excel文件创建与txt不一样，txt用open可以创建
        wb=Workbook()
        wb.save(copyexcel)
        # 创建excel对象
        copybook = xlrd.open_workbook(copyexcel)
        # 创建可写excel对象
        wcopy_book = copy(copybook)
        copy_book_sheet = wcopy_book.get_sheet(0)
        # 把读出来的excel内容写到副本
        for r, values in zip(range(nrows), content):
            for c, v in zip(range(cols), values):
                copy_book_sheet.write(r, c, v)
        wcopy_book.save(copyexcel)
        return copyexcel
'''
    # 复制用例，结果写入副本，目的是用例不用每次运行都要关闭，避免用例打开，写入结果出错
    # dirpath：目录, filepath：文件具体地址, filepathnext：新的目录名
    def copy_excel(self, dirpath, filepath, filepathnext='备份',iscopy=None):
        # 在当前目录下创建一个新的目录用于存入复制得到的EXCEL
        newfilepath = os.path.join(dirpath, filepathnext)
        # 创建目录
        if not os.path.exists(newfilepath):
            os.mkdir(newfilepath)
            # print('输出创建的目录', newfilepath)
        else:
            # print('不需要创建访目录')
            pass
        # 把原文件名取出来，以用于创建副件名
        copyexcel = os.path.join(newfilepath, os.path.basename(filepath).split('.')[0] + '_副本' + '.xlsx')

        try:
            # 创建空excel,excel文件创建与txt不一样，txt用open可以创建
            if iscopy is None:
                # exists 判断不稳定
                if os.path.exists(copyexcel):
                # if os.path.isfile(copyexcel):
                    print('文件已经存在，不需要创建')
                else:
                    wb=xlsxwriter.Workbook(copyexcel)
                    # wb = xlwt.Workbook()
                    # sheet = wb.add_sheet('Sheet1')
                    # print(copyexcel)
                    # wb.save(copyexcel)
                    # wb.close()
                    print('创建副本成功')
                # 复制文件
                shutil.copy(filepath, copyexcel)
                print('文件复制成功')

            else:
                if os.path.exists(copyexcel):
                    print('不复制，副本存在')
                else:
                    print('不复制，但副本不存在')
                    return 'path_not_exits'

        except Exception as e:
            print('复制文件出错！',e)

        return copyexcel

    # 把结果写回excel
    #当文件打开时，运行方法会报错（或者数据写不进去！！！！！），不允许操作，所以运作时一定要关闭文件，另外重新运作，新插入数据会覆盖原有的
    # row:用例所在行数, actualresult：实际用例返回结果, result：对比断言后的结果, user：登录系统的账号, casefile：回写结果的用例文件
    def write_result(self, row, actualresult,result, casefile):  # 文件不用写全路径，只写文件名只可
        self.lock.acquire()
        case_path=os.path.join(CASEPATH, casefile)
        print(case_path)
        book = xlrd.open_workbook(case_path)  # 创建一个excel操作对象
        book2 =copy(book) #复制book对象  #管道作用:利用xlutils.copy函数，将xlrd.Book转为xlwt.Workbook，再用xlwt模块进行存储
        sheet = book2.get_sheet(0)
        # 创建一个sheet操作实例,读取的是第一个excel ,#通过get_sheet()获取的sheet有write()方法
        # 红色
        # styleBlueBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
        style = XFStyle()     # 格式信息
        font = xlwt.Font()  # 字体基本设置
        # 红色
        font.colour_index=0xff
        style.font=font
        sheet.write(row, 11, actualresult)
        # 回写时，如果是失败，则颜色标成红色
        # if result=='0':
        #     # sheet.write(row, 12, result,styleBlueBkg)
        #     sheet.write(row, 12, result,style)
        # else:
        #     sheet.write(row, 12, result)
        sheet.write(row, 12, result)
        # sheet.write(row, 13, user)
        book2.save(case_path)
        print('结果写入保存成功！')
        self.lock.release()
        # sheet2=book.sheet_by_index(0)
        # print(sheet2.cell_value(row,12))

    # 订单导入，批量更新订单号，如果传入copydata则是指定了要创建的行数，方便快速做批量的订单,isparam表示是否要参数化
    def order_upload(self,filepath,copydata=None,isparam=None):  # filepath路径，copydata：订单总行数
        # self.lock.acquire()
        book = xlrd.open_workbook(filepath)  # 创建一个excel操作对象
        newname=filepath
        # 第一个表格
        sheet1= book.sheet_by_index(0)
        # 第二个表格
        # sheet12= book.sheet_by_index(1)
        # 第一个表格行数，包含表头
        nr1=sheet1.nrows
        print(nr1)
        # 如果获取到只有一行，说明只有表头，没有要导入的数据
        if nr1 ==1:
            return nr1,newname
        else:
            pass
        # 第一个表格列数
        nc1=sheet1.ncols
        # 第二个表格行数
        # nr12=sheet12.nrows
        book2 =copy(book) #复制book对象  #管道作用:利用xlutils.copy函数，将xlrd.Book转为xlwt.Workbook，再用xlwt模块进行存储
        # 第一个表格
        rsheet1 = book2.get_sheet(0)  # 创建一个sheet操作实例,读取的是第一个excel ,#通过get_sheet()获取的sheet有write()方法
        tm = int(time.strftime('%m%d%H%M%S',time.localtime(time.time())))
        # 若copydata为 None，则表示不需要复制数据
        if copydata is None:
            pass
        # 当输入的copydata大于当前行数才进行复制粘贴，其他情况下直接通过跳到后续操作。nr1-1目的是排除第一行表头
        elif copydata > nr1-1:
            # 获取第二行（第一行是表头），以此来复制粘贴
            orderlist=sheet1.row_values(1)
            print(orderlist)
            # 第一层循环是要重复的行数
            for i in rangeyield2((copydata-nr1)+1):
                # 第二层循环是复制第一行的数据粘贴到每一行，需要逐一写到每一列
                for j,k in zip(orderlist,rangeyield2(nc1)):
                    rsheet1.write(nr1+i,k,j)
            # 更新最新行数
            nr1 = copydata+1

        elif copydata< 0 or copydata ==0 :
            print('输入的不为none，小于等于0,不复制')
            return nr1,newname
        # copydata比文件行数小，为避免修改导入的订单比要修改的订单多，所以修改导入的模板要用新文件，导入订单的不用新文件
        else:
            if isparam is not None:
                # 把第一数据保存到列表
                orderlist = sheet1.row_values(1)
                orderlist0 = sheet1.row_values(0)
                # 清除所有数据,delete_cols 删除不生效，rsheet1.write(i, j, '')，rsheet1.write(i, j, None) 删除不干净，
                # 所以只能新建文件并复制到新文件，在上传完后删除文件
                # 判断文件是否已经存在，若存在则先删除
                newname = os.path.join(os.path.dirname(filepath),
                                       os.path.basename(filepath).split('.')[0] + '_backup' + '.xlsx')
                # 判断文件存在则先删除
                if os.path.exists(newname):
                    os.remove(newname)
                else:
                    pass
                # 创建文件
                wb = xlsxwriter.Workbook(newname)
                # wb.add_worksheet('new')
                wb.close()
                book3 = xlrd.open_workbook(newname)
                book4 = copy(book3)
                sheet4 = book4.get_sheet(0)
                # 把表头复制进去
                for j, k in zip(orderlist0, rangeyield2(nc1)):
                    sheet4.write(0, k, j)
                # # 拿第一行进行复制
                for i in rangeyield2(copydata):
                    # 第二层循环是复制第一行的数据粘贴到每一行，需要逐一写到每一列
                    for j, k in zip(orderlist, rangeyield2(nc1)):
                        sheet4.write(i + 1, k, j)
                # 更新最新行数
                nr1 = copydata + 1
                print(121121112)
                # 保存
                book4.save(newname)
            # 判断是订单导入，则不需要处理
            else:
                pass
        # 判断是否进入参数化,目的是修改导入模板不需要自动参数化
        if isparam is None:
            # 第二个表格
            sheet12 = book.sheet_by_index(1)
            # 第二个表格行数
            nr12 = sheet12.nrows
            # 第二个表格
            rsheet2 = book2.get_sheet(1)  # 创建一个sheet操作实例,读取的是第一个excel ,#通过get_sheet()获取的sheet有write()方法
            # 主表订单号参数化
            # 从第二行开始，第一行是表头
            for row, value in zip([row for row in rangeyield(1, nr1)], [value for value in rangeyield(tm, tm + nr1)]):
                # print(row,value)
                rsheet1.write(row, 3, value)
            # 第二页第一行是列名，不等于一行证明有SKU，则订单号进行参数化
            if nr12 != 1:
                for row, value in zip([row for row in rangeyield(1, nr12, 1)],
                                      [value for value in rangeyield(tm, tm + nr12 - 1)]):
                    rsheet2.write(row, 2, value)
            else:
                print('第二个表格没有明细')
        else:
            return nr1,newname

        book2.save(filepath)
        # self.lock.release()
        print('结果写入保存成功！')
        print('打印返回的row',nr1)
        return nr1,newname

    # 删除文件，目前的应用场景是删除report中的文件，以便运行完获取新增的报告文件发送邮件，所以多进程运行时，最好也是这个目录下执行
    # dirpath：目录，filepath：具体文件地址
    def delete_file(self,dirpath=REPORTPATH, filepath=None):
        try:
            # 若输入具体路径，删除具体路径的文件
            if filepath is not None:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    return
                else:
                    print('输入的具体路径不存在，不需要删除')
            else:
                files = os.listdir(dirpath)
                # 长度大于0，说明有文件或目录可删除
                if len(files) > 0:
                    # for 循环中的代码，有时候会误判断没有找到文件，路径无效
                    for f in files:
                        # print(f)
                        # 删除文件和删除目录的方式不一样，所以需要分开操作
                        if os.path.isfile(os.path.join(dirpath, f)):
                            if os.path.exists(os.path.join(dirpath, f)):
                                os.remove(os.path.join(dirpath, f))
                            else:
                                print('文件不存在')


                        else:
                            if os.path.exists(os.path.join(dirpath, f)):
                                # 删除目录，空目录，有文件的目录
                                shutil.rmtree(os.path.join(dirpath, f))
                            else:
                                print('文件不存在')

                        print('删除文件结束')

                    # name=os.path.split(dirpath)[1]
                    # 整个文件夹删除; 由于bat方式要切换到此目录运行，所以删除会出错
                    # shutil.rmtree(dirpath)
                    # print('文件夹及其下内容全部删除')
                    # time.sleep(4)
                    # # 再创建一个空的文件夹
                    # os.mkdir(dirpath)
                    # print('新文件夹创建成功')

                else:
                    print('目录下没有文件，不需要删除')


        except Exception as e:
            print(e)
    # 压缩文件夹,src_dir:要压缩目录的路径
    def zip_file(self,src_dir):
        zip_name = src_dir +'.zip'
        # 判断压缩文件已经存在，则先删除
        if os.path.exists(zip_name):
            os.remove(zip_name)
        else:
            pass
        z = zipfile.ZipFile(zip_name,'w',zipfile.ZIP_DEFLATED)
        for dirpath, dirnames, filenames in os.walk(src_dir):
            fpath = dirpath.replace(src_dir,'')
            fpath = fpath and fpath + os.sep or ''
            for filename in filenames:
                z.write(os.path.join(dirpath, filename),fpath+filename)
                # print ('==压缩成功==')
        z.close()
        return zip_name

    # filepath:文件路径, col ：列, row：行,用列表，允许输入两个元素，第一与第二个分别表示要获取的开始与结束值
    # row 要求必须输入开始值，col要求必须输入开始值
    def excel_operat(self,filepath, col=[], row=[], *args):
        try:
            book = xlrd.open_workbook(filepath)  # 创建一个excel操作对象
            # print(os.path.join(CASEPATH, file))
            sheet = book.sheet_by_index(0)  # 创建一个sheet操作实例,读取的是第一个excel，如果需要也可以参数化
            nnrows = sheet.nrows  # 获取行数,也就是最大行数
            nncols = sheet.ncols  # 获取行数,也就是最大行数
            # 默认最小行数是0
            startrow=0
            startcol=0

            if len(row)>0:
                if len(row)>2:
                    print('输入row列表长度大于2，不合法'.format(row))
                    return
                else:
                    if len(row) == 2:
                        startrow=row[0]
                        nnrows = row[1]
                    # 等于1，只输入开始
                    else:
                        startrow = row[0]
            else:
                print('输入row列表长度小于0，不合法'.format(row))
                return

            if len(col)>0:
                if len(col)>2:
                    print('输入col列表长度大于2，不合法'.format(col))
                    return
                else:
                    if len(col) == 2:
                        startcol=col[0]
                        nncols = col[1]
                    # 长度等于1，开始值
                    else:
                        startcol=col[0]
            else:
                print('输入row列表长度小于0，不合法'.format(row))
                return
            # 加入列表并返回
            mylist = []
            for i in range(startrow,nnrows):
                mylist.append(sheet.row_values(i)[startcol:nncols])
        except Exception as e:
            print(e)
        # print(startcol,startrow,type(nncols),type(nnrows))
        return mylist

    # 写入数据
    # row 是列表，可以有两个值，col，是列表，最多有一个值，目的是只针对某列操作，valuelist 是列表，表示要写入的值的列表，filepath，表示文件名
    def write_excel(self, row=[], col=[],valuelist=[], filepath=None):  # 文件不用写全路径，只写文件名只可
        # self.lock.acquire()
        book = xlrd.open_workbook(filepath)  # 创建一个excel操作对象
        sheet = book.sheet_by_index(0)  # 创建一个sheet操作实例,读取的是第一个excel，如果需要也可以参数化
        nnrows = sheet.nrows  # 获取行数,也就是最大行数
        nncols = sheet.ncols  # 获取行数,也就是最大行数
        # 默认最小行数是0
        startrow = 0
        startcol = 0
        if len(row) > 0:
            if len(row) > 2:
                print('输入row列表长度大于2，不合法'.format(row))
                return
            else:
                if len(row) == 2:
                    startrow = row[0]
                    nnrows = row[1]
                # 等于1，只输入开始
                else:
                    startrow = row[0]
        else:
            print('输入row列表长度小于0，不合法'.format(row))
            return

        if len(col) > 0:
            if len(col) > 2 or  len(col) ==2:
                print('输入col列表长度大于等于2，不合法'.format(col))
                return
            else:
                startcol = col[0]
        else:
            print('输入row列表长度小于0，不合法'.format(row))
            return
        book2 = copy(book)  # 复制book对象  #管道作用:利用xlutils.copy函数，将xlrd.Book转为xlwt.Workbook，再用xlwt模块进行存储
        sheet2 = book2.get_sheet(0)
        # 写入
        for row,value in zip(range(startrow, nnrows),valuelist):
            sheet2.write(row, startcol, value)
        book2.save(filepath)
        print('结果写入保存成功！')
        # self.lock.release()



#方法调试
if __name__ == '__main__':
    do=DataOperate()
    # do.order_upload(filepath='D:\Personal\moshuangyou\Desktop\EDI订单批量导入.xlsx',copydata=5000)
    # do.order_upload(filepath='D:\Personal\moshuangyou\Desktop\订单批量0923 - 副本.xlsx',copydata=3000)
    # do.order_upload(filepath='D:\Personal\moshuangyou\Desktop\import0525.xlsx',copydata=5000)
    # do.copy_excel(dirpath='E:\\software\\python3.5.1\\UiAuto\\case',filepath='E:\\software\\python3.5.1\\UiAuto\\case\\POcase.xlsx')
    do.delete_file()
